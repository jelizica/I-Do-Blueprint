#!/usr/bin/env python3
"""
Convert CSV sample files to XLSX format for Excel import testing.

This script converts CSV sample files to XLSX format with proper formatting
for testing the Excel import functionality in the I Do Blueprint app.

Usage:
    python3 Scripts/convert_csv_to_xlsx.py

Requirements:
    pip3 install openpyxl

Author: I Do Blueprint Development Team
Last Updated: December 2025
"""

import csv
import sys
from pathlib import Path
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("❌ Error: openpyxl is not installed.")
    print("📦 Please install it with: pip3 install openpyxl")
    sys.exit(1)


def convert_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """
    Convert a CSV file to XLSX format with formatting.
    
    Args:
        csv_path: Path to the source CSV file
        xlsx_path: Path where the XLSX file will be created
        
    Raises:
        FileNotFoundError: If the CSV file doesn't exist
        PermissionError: If unable to write the XLSX file
        csv.Error: If the CSV file is malformed
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # Read CSV and write to Excel
        with open(csv_path, 'r', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            
            for row_idx, row in enumerate(reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Format header row
                    if row_idx == 1:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(
                            start_color="D3D3D3",
                            end_color="D3D3D3",
                            fill_type="solid"
                        )
                        cell.alignment = Alignment(
                            horizontal="left",
                            vertical="center"
                        )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    # Skip cells that can't be converted to string
                    pass
            
            # Cap at 50 characters for readability
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        wb.save(xlsx_path)
        print(f"✅ Created: {xlsx_path.name}")
        
    except csv.Error as e:
        print(f"❌ Error reading CSV file: {e}")
        raise
    except PermissionError as e:
        print(f"❌ Error writing XLSX file: {e}")
        raise
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
        raise


def main() -> int:
    """
    Main entry point for the script.
    
    Returns:
        0 on success, 1 on error
    """
    try:
        # Get the project root directory
        script_dir = Path(__file__).parent
        project_root = script_dir.parent
        resources_dir = project_root / "I Do Blueprint" / "Resources"
        
        # Verify resources directory exists
        if not resources_dir.exists():
            print(f"❌ Error: Resources directory not found: {resources_dir}")
            print("💡 Make sure you're running this script from the project root")
            return 1
        
        # Define file paths
        files_to_convert = [
            ("SampleGuestList.csv", "SampleGuestList.xlsx"),
            ("SampleVendorList.csv", "SampleVendorList.xlsx"),
        ]
        
        # Check if all CSV files exist
        missing_files = []
        for csv_name, _ in files_to_convert:
            csv_path = resources_dir / csv_name
            if not csv_path.exists():
                missing_files.append(csv_name)
        
        if missing_files:
            print("❌ Error: Missing CSV files:")
            for filename in missing_files:
                print(f"  - {filename}")
            print(f"\n💡 Expected location: {resources_dir}")
            return 1
        
        print("🔄 Converting CSV files to XLSX format...\n")
        
        # Convert all files
        converted_files = []
        for csv_name, xlsx_name in files_to_convert:
            csv_path = resources_dir / csv_name
            xlsx_path = resources_dir / xlsx_name
            
            try:
                convert_csv_to_xlsx(csv_path, xlsx_path)
                converted_files.append(xlsx_path)
            except Exception as e:
                print(f"❌ Failed to convert {csv_name}: {e}")
                return 1
        
        print("\n✅ Conversion complete!")
        print(f"\n📁 Created {len(converted_files)} file(s):")
        for xlsx_path in converted_files:
            print(f"  - {xlsx_path.name}")
        
        return 0
        
    except KeyboardInterrupt:
        print("\n\n⚠️  Conversion cancelled by user")
        return 1
    except Exception as e:
        print(f"\n❌ Unexpected error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
